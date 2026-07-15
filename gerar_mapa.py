import calendar
from datetime import date, timedelta
import os
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import FormulaRule
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from estilos import *
from openpyxl.worksheet.datavalidation import DataValidation

def gerar_mapa(ficheiro):

    # Ler folha "Ferias"
    ferias = pd.read_excel(
        ficheiro,
        sheet_name="Ferias"
    )

    # Renomear colunas
    ferias = ferias.rename(columns={
        "Colaborador ID": "ID",
        "Data de início": "Data Inicio",
        "Data de fim": "Data Fim"
    })

    # Converter datas
    ferias["Data Inicio"] = pd.to_datetime(
        ferias["Data Inicio"],
        dayfirst=True
    )

    ferias["Data Fim"] = pd.to_datetime(
        ferias["Data Fim"],
        dayfirst=True
    )

    anos = ferias["Data Inicio"].dt.year.unique()

    if len(anos) != 1:
        raise ValueError(
            "A folha 'Ferias' contém registos de mais do que um ano."
        )

    ano = anos[0]

    # Criar workbook
    wb = Workbook()

    ws = wb.active
    ws.title = "Mapa"

    # Construir o cabeçalho
    mapa_datas = criar_cabecalho(ws, ano)

    # Escrever os meses
    escrever_meses(ws, ano)

    # Escrever colaboradores e pintar férias
    preencher_colaboradores(
        ws,
        ferias,
        mapa_datas
    )

    ultima_coluna = max(mapa_datas.values())

    ultima_linha = (
        ferias["Nome"]
        .drop_duplicates()
        .shape[0]
        + 4
    )

    adicionar_lista_ferias(
        ws,
        ultima_coluna,
        ultima_linha
    )
    
    # Guardar

    os.makedirs("output", exist_ok=True)

    caminho = os.path.join(
        "output",
        f"Mapa_Ferias_{ano}.xlsx"
    )

    wb.save(caminho)

    return caminho

def preencher_colaboradores(ws, ferias, mapa_datas):

    # Colaboradores únicos ordenados por nome
    colaboradores = (
        ferias[["ID", "Nome"]]
        .drop_duplicates()
        .sort_values("Nome")
        .reset_index(drop=True)
    )

    linha = 5

    linhas_colaboradores = {}

    # Escrever nomes
    for _, colaborador in colaboradores.iterrows():

        linhas_colaboradores[colaborador["ID"]] = linha

        celula = ws.cell(row=linha, column=1)

        celula.value = colaborador["Nome"]
        celula.fill = COR_NOME
        celula.font = FONTE_NORMAL
        celula.alignment = CENTRO
        celula.border = BORDA

        # Pintar toda a linha (365/366 dias)
        for data, coluna in mapa_datas.items():

            c = ws.cell(row=linha, column=coluna)

            if data.weekday() >= 5:
                c.fill = COR_FIMSEMANA
            else:
                c.fill = COR_TRABALHO

            c.border = BORDA

        linha += 1


    # Pintar férias
    for _, registo in ferias.iterrows():

        linha = linhas_colaboradores[registo["ID"]]

        dia = registo["Data Inicio"].date()
        fim = registo["Data Fim"].date()

        while dia <= fim:

            coluna = mapa_datas[dia]

            celula = ws.cell(row=linha, column=coluna)

            celula.value = "F"
            celula.border = BORDA
            celula.alignment = CENTRO

            dia += timedelta(days=1)
    
    ultima_coluna = get_column_letter(max(mapa_datas.values()))

    for linha in linhas_colaboradores.values():

        celula = ws.cell(row=linha, column=2)

        celula.value = (
            f'=COUNTIF(C{linha}:{ultima_coluna}{linha},"F")'
        )

        celula.font = FONTE_BOLD
        celula.alignment = CENTRO
        celula.border = BORDA


def criar_cabecalho(ws, ano):

    mapa_datas = {}

    ws["A4"] = "Colaborador"
    ws["B4"] = "Dias de Férias"

    for cel in ("A4", "B4"):
        ws[cel].fill = COR_NOME
        ws[cel].font = FONTE_BOLD
        ws[cel].alignment = CENTRO
        ws[cel].border = BORDA

    # ---------- Título ----------
    ws["A1"] = f"MAPA ANUAL DE FÉRIAS {ano}"
    ws["A1"].font = FONTE_TITULO
    ws["A1"].alignment = CENTRO

    dias_semana = [
        "Seg",
        "Ter",
        "Qua",
        "Qui",
        "Sex",
        "Sáb",
        "Dom"
    ]

    data = date(ano, 1, 1)
    coluna = 3

    while data.year == ano:

        mapa_datas[data] = coluna

        c_dia = ws.cell(row=3, column=coluna)
        c_num = ws.cell(row=4, column=coluna)

        c_dia.value = dias_semana[data.weekday()]
        c_num.value = data.day

        if data.weekday() >= 5:
            fill = COR_FIMSEMANA
        else:
            fill = COR_TRABALHO

        c_dia.fill = fill
        c_num.fill = fill

        c_dia.font = FONTE_BOLD
        c_num.font = FONTE_NORMAL

        c_dia.alignment = CENTRO
        c_num.alignment = CENTRO

        c_dia.border = BORDA
        c_num.border = BORDA

        coluna += 1
        data += timedelta(days=1)

    # Unir o título até à última coluna
    ultima_coluna = coluna - 1

    ws.merge_cells(
        f"A1:{get_column_letter(ultima_coluna)}1"
    )

    # Larguras
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10

    for c in range(3, coluna):
        ws.column_dimensions[get_column_letter(c)].width = 4

    # Congelar painéis
    ws.freeze_panes = "C5"

    # Filtro
    ws.auto_filter.ref = f"A4:{get_column_letter(ultima_coluna)}4"

    verde = PatternFill(
        fill_type="solid",
        start_color="78ED74",
        end_color="78ED74"
    )

    fonte_verde = Font(color="78ED74")

    ultima_coluna = get_column_letter(coluna - 1)

    regra = FormulaRule(
        formula=['C5="F"'],
        fill=verde,
        font=fonte_verde
    )

    ws.conditional_formatting.add(
        f"C5:{ultima_coluna}500",
        regra
    )   

    return mapa_datas


def escrever_meses(ws, ano):

    coluna_inicio = 3

    for mes in range(1, 13):

        dias_mes = calendar.monthrange(ano, mes)[1]

        coluna_fim = coluna_inicio + dias_mes - 1

        ws.merge_cells(
            start_row=2,
            start_column=coluna_inicio,
            end_row=2,
            end_column=coluna_fim
        )

        MESES = [
            "JANEIRO",
            "FEVEREIRO",
            "MARÇO",
            "ABRIL",
            "MAIO",
            "JUNHO",
            "JULHO",
            "AGOSTO",
            "SETEMBRO",
            "OUTUBRO",
            "NOVEMBRO",
            "DEZEMBRO"
        ]

        celula = ws.cell(row=2, column=coluna_inicio)
        celula.value = MESES[mes - 1]

        celula.fill = COR_MES
        celula.font = FONTE_BOLD
        celula.alignment = CENTRO
        celula.border = BORDA

        # Pintar todas as células do mês
        for c in range(coluna_inicio, coluna_fim + 1):
            ws.cell(row=2, column=c).fill = COR_MES
            ws.cell(row=2, column=c).border = BORDA

        coluna_inicio = coluna_fim + 1

def adicionar_lista_ferias(ws, ultima_coluna, ultima_linha):

    dv = DataValidation(
        type="list",
        formula1='"F"',
        allow_blank=True
    )

    dv.prompt = "Selecione F para marcar férias."
    dv.error = "Valor inválido."

    ws.add_data_validation(dv)

    intervalo = (
        f"C5:"
        f"{get_column_letter(ultima_coluna)}{ultima_linha}"
    )

    dv.add(intervalo)