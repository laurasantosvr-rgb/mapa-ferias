from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Cores
COR_NOME = PatternFill("solid", fgColor="B6F5FA")        # Azul claro
COR_MES = PatternFill("solid", fgColor="5B9BD5")         # Azul
COR_DIA = PatternFill("solid", fgColor="07DEFA")         # Azul escuro
COR_TRABALHO = PatternFill("solid", fgColor="F2F2F2")    # Cinzento claro
COR_FIMSEMANA = PatternFill("solid", fgColor="A6A6A6")   # Cinzento escuro
COR_FERIAS = PatternFill("solid", fgColor="78ED74")      # Verde

# Fontes
FONTE_TITULO = Font(size=16, bold=True)
FONTE_NORMAL = Font(size=10)
FONTE_BOLD = Font(size=10, bold=True)

# Alinhamento
CENTRO = Alignment(horizontal="center", vertical="center")

# Bordas
BORDA = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)